"""
Nützliche Hilfs-Funktionen 
für Datenhandling bei Vektorparametern und vektorwertigen Variablen
Version 2.2 vom 17. März 2026
@author: Kai Hege Becker

Funktionen:
    output(): Gibt Variable in Konsole aus
    vectorparam(): Liest Vektorparameter aus Python-Daten-File ein
    convert_celladdress(): Konvertiert xls-Zellenadresse (Hilfsfuntion)
    open_sheet(): Öffnet xls-File zum Lesen und Schreiben
    save_sheet(): Speichert xls-File nach Schreiben
    read_scalar(): Liest den Wert eines Skalars aus xls-File
    read_index(): Liest die Indizes einer Indexmenge aus xls-File
    read_table(): Liest die Werte eines Vektorparameters aus xls-File
    write_scalar(): Schreibt einen Skalar mit Überschrift in xls-File
    write_thead(): Schreibt den Kopf einer xls-Tabelle
    write_tbody(): Schreibt xls-Tabelle mit den Werten einer Variablen
    write_lhead(): Schreibt Kopf einer Liste in ein xls-File
    write_lbody(): Gibt Liste mit den Werten einer Variablen in xls-File aus
    out_rc(): Gibt die reduzierten Kosten einer Vektorvariablen aus
    out_sl(): Gibt die Schlupfe eines Nebenbedingungstyps aus
    out_dv(): Gibt die Dualwerte eines Nebenbedingungstyps aus
    saOFC(): Gibt Sensitivitätsanalyse von ZFK einer Vektorvariablen aus
    saRHS(): Gibt Sensitivitätsanalyse von rechter Seite eines NB-Typs aus
    
"""

import itertools
import re
from openpyxl import load_workbook

   
def output(name, var, *args):
    """
    Printet diejenigen Values eines Dictionarys mit Tupeln als Keys, 
    die (praktisch) ungleich Null sind. 

    Parameters
    ----------
    name : string
        Auszugebender Text (z.B. Name der Variablen)
    var : dict
        Dictionary mit Tupeln als Keys
    *args : Iterable
        Indexmengen der Keys des mehrdimensionalen Dictionarys

    Returns
    -------
        None.
    """
    print()
    print("******************************")
    if len(args) > 1:
        for i in itertools.product(*args):
            if var[i] >= 1e-10 or var[i] <= -1e-10:
                print(name, i, " = ", var[i])
    else:
        (iterable,) = args
        for i in iterable:
            if var[i] >= 1e-10 or var[i] <= -1e-10:
                print(name, i, " = ", var[i] )
            

def vectorparam(var, *args) -> dict:
    """
    Erzeugt Vektorparameter für Optimierungsmodell. 
    Genauer:
    Erzeugt aus einem Dictionary, das 1D- oder 2D-Listen als Values hat, 
    ein Dictionary, das die Elemente der Listen als Values hat.
    Erzeugt auch aus einer 1D- oder 2D-Liste ein Dictionary,
    das die Elemente der Liste als Values hat.

    Parameters
    ----------
    var : Dict[List]
        Ausgangs-Dictionary mit Tupeln (mind. 2D) als Keys 
        und 1D- oder 2D-Listen als Values
        oder nur eine 1D- oder 2D-Ausgangs-Liste.
    *args : Iterable
        Iterablen, deren Kreuzprodukt (als Tupel) die Keys des zu erzeugenden
        Dictionarys darstellen.

    Returns
    -------
    new_dict : Dict
        Dictionary, das als Values die Einträge der Listen des
        Ausgangs-Dictionarys hat und Tupel aller Iterablen aus *arg als Keys.
    """
    new_dict = {}
    if isinstance(var, list):       #Ausgangsdaten sind nur Liste
        if not isinstance(var[0], list):     #1D-Liste
            (iterable,) = args                              #Indexmenge 
            iterator = iter(var)
            for i in iterable:
                new_dict[i] = next(iterator)
        else:                                #2D-Liste
            iterator = itertools.chain.from_iterable(var)   
            for i in itertools.product(*args):              #Indexmenge
                new_dict[i] = next(iterator)
                
    else:                           #Ausgangsdaten sind mind. 2D-Dictionary
        n = len(next(iter(var)))    #Dimension der Key-Tupels
        if len(args[n:]) == 1:               #mit 1D-Liste als Values
            (iterable,) = args[n:]                          #Indexmenge
            for i in itertools.product(*args[0:n]):         #weiter Indexmenge
                iterator = iter(var[i])
                for j in iterable:
                    new_dict[i+(j,)] = next(iterator)
        else:                                #mit 2D-Liste als Values
            for i in itertools.product(*args[0:n]):         #Indexmenge
                iterator = itertools.chain.from_iterable(var[i])
                for j in itertools.product(*args[n:]):      #weiter Indexmenge
                    new_dict[i+j] = next(iterator)
        
    return new_dict

def convert_celladdress(address) -> (int, int):
    """
    Kovertiert eine Zellen-Adresse in Excel-Format (gegeben als string)
    in eine Zeilen- und eine Spaltennummer (Zählung beginnend von 1)
    """
    twostrings = re.split('(\d+)', address)
    row = int(twostrings[1])
    loc = len(twostrings[0])
    column = 0
    for i in range(0, loc):
        place = loc - i - 1
        letternumber = ord(twostrings[0][i]) - 64
        column = column + letternumber*pow(26, place)
    return row, column

def open_sheet(filename, sheetnumber):
    """
    Öffnet ein Sheet in einem Excel-File 
    und gibt Referenz auf Workbook und Sheet zurück
    """
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[sheetnumber]
    return workbook, sheet

def save_sheet(workbook, filename):
    """
    Speichert das angegebene Workbook-Objekt unter dem angegebenen Dateinamen
    """
    workbook.save(filename)
    
def read_scalar(sheet, startcell):
    """
    Liest einen Skalar aus einer Excel-Tabelle

    Parameters
    ----------
    sheet : object
        Referenz des zu lesenden Worksheets.
    startcell : string
        Zelle, aus der der Wert des Skalars gelesen wird.

    Returns
    -------    
    Wert, der aus der Zelle gelesen wird.
    """
    r, c = convert_celladdress(startcell)
    return sheet.cell(row=r, column=c).value

def read_index(sheet, startcell, direction, length):
    """
    Liest Indexmenge aus einer Excel-Tabelle und speichert sie als Liste

    Parameters
    ----------
    sheet : object
        Referenz der zu lesenden Excel-Tabelle.
    startcell : string
        Zelle, in der die Indices beginnen.
    direction: string
        Ausrichtung der Indexmenge in Tabelle ("vertical" oder "horizontal")    
    length : int
        Länge der zulesenden Zeile oder Spalte.

    Returns
    -------
    indexset: list(!) 
        Gelesene Indexmenge.

    """
    r, c = convert_celladdress(startcell)
    indexset = []
    if direction == "horizontal":
        for i in range(length):
            indexset.append(sheet.cell(row=r, column=c+i).value)
    if direction == "vertical":
        for i in range(length):
            indexset.append(sheet.cell(row=r+i, column=c).value)
    return indexset

def read_table(sheet, startcell, *args):
    """
    Liest ein Dictionary aus einer Excel-Tabelle,
    das als Keys Tupel hat, die das Kreuzprodukt von Indexmengen sind.

    Parameters
    ----------
    sheet : object
        Referenz der zu lesenden Excel-Tabelle.
    startcell : string
        Zelle, in der die Tabellenwerte beginnen.
    *agrs: iterables, 
        Indexmengen, deren Kreuzprodukt die Keys darstellen    

    Returns
    -------
    dict: dictionary 
        Gelesenes Dictionary.

    """
    r, c = convert_celladdress(startcell)
    dictionary = {}
    if len(args) == 1:                  # eine Indexmenge
        (iterable,) = args
        for i in range(len(iterable)):
            key = sheet.cell(row=r+i, column=c-1).value
            dictionary[key] = sheet.cell(row=r+i, column=c).value
    else:                               # mehrere Indexmengen
        # Berechne die Anzahl der Zeilen der Tabelle
        numrow = 1
        for i in args[:-1]:
            numrow = numrow * len(i)
        # Berechne die Anzahl der Spalten der Tabelle
        numcol = len(args[-1])
        # Lies Tabelle
        for i in range(numrow):
            keyrows = ()
            for j in range(len(args[:-1])):
                rowkey =  sheet.cell(row=r+i, column=c-len(args[:-1])+j).value
                keyrows = keyrows + (rowkey,)
            for k in range(numcol):
                colkey = sheet.cell(row=r-1, column=c+k).value
                key = keyrows + (colkey,)
                dictionary[key] = sheet.cell(row=r+i, column=c+k).value
    return dictionary

def write_scalar(sheet, startcell, var, name):
    """
    Schreibt die Werte einer Variable in eine Excel-Tabelle

    Parameters
    ----------
    sheet : object
        Referenz des zu beschreibenden Worksheets.
    startcell : string
        Erste Zelle der Tabelle, enthält die Überschrift.
        Die Variable wird in die Zeile darunter geschrieben.
    var : Name der auszugebenden Variablen.
    name : string
        Auszugebende Überschrift

    Returns
    -------
    None.

    """
    r, c = convert_celladdress(startcell)
    sheet.cell(row=r, column=c).value=name
    sheet.cell(row=r+1, column=c).value=var

def write_thead(sheet, startcell, var, *args):
    """
    Schreibt die Kopfzeilen für eine Excel-Tabelle
    für eine Variable, die danach mit write_tbody() ausgegeben wird.

    Parameters
    ----------
    sheet : object
        Referenz des zu beschreibenden Worksheets.
    startcell : string
        Erste Zelle der Tabelle. Ab hier wird die Kopfzeile geschrieben.
    var : string,
        Name der in der Tabelle auszugebenden Variablen.
    *args : string, 
        Namen der Indexmengen der auszugebenden Variablen.

    Returns
    -------
    None.

    """

    r, c = convert_celladdress(startcell)
    # Schreibe Tabellenkopf
    if len(args) == 1:               # eindimensionale Variable
        i = 0
        for j in args:
            sheet.cell(row=r, column=c+i).value=j
            i = i + 1
        sheet.cell(row=r, column=c+i).value=var
    else:                            # mehrdimensionale Variable    
        # Name der Variablen
        sheet.cell(row=r, column=c).value=var
        # Name der Indexmengen bis auf die letzte
        i = 0
        for j in args[:-1]:
            sheet.cell(row=r+1, column=c+i).value=j
            i = i + 1
        # Name der letzten Indexmenge    
        sheet.cell(row=r, column=c+i).value=args[-1]  
        
    
def write_tbody(sheet, startcell, var, *args):
    """
    Schreibt die Werte einer Variable in eine Excel-Tabelle
    Im Gegensatz zur write_lbody() werden hier auch Variablenwerte
    ausgegeben, die (praktisch) Null sind.

    Parameters
    ----------
    sheet : object
        Referenz des zu beschreibenden Worksheets.
    startcell : string
        Erste Zelle der Tabelle, dort beginnt der Tabellenkopf,
        der mit write_thead() geschrieben wurde.
        Der Tabellenkörper beginnt darunter.
    var : Dictionary mit Tupeln als Keys (Name der auszugebenden Variablen).
    *args : iterable, 
        Indexmengen der Keys des mehrdimensionalen Dictionarys.

    Returns
    -------
    None.

    """

    r, c = convert_celladdress(startcell)
    # Schreibe Rest der Tabelle
    if len(args) > 1:                # falls Indexmenge mehrdimensional ist
        # Indices der letzten Indexmenge
        m = len(args[:-1])
        for l in args[-1]:
            sheet.cell(row=r+1, column=c+m).value=l
            m = m + 1
        # Gehe durch die Zeilen der Tabelle
        i = 2
        for j in itertools.product(*args[:-1]):
            # Indices der ersten Indexmengen für eine Zeile
            for k in range(0, len(args[:-1])):
                sheet.cell(row=r+i, column=c+k).value=j[k]
            # Werte der Variablen für eine Zeile   
            m = 1
            for l in args[-1]:
                key = j + (l,)
                sheet.cell(row=r+i, column=c+k+m).value=var[key]
                m = m + 1
            i = i + 1 
    else:
        (iterable,) = args           # falls Indexmenge eindimensional ist
        i = 1                       
        for j in iterable:
            sheet.cell(row=r+i, column=c).value=j
            sheet.cell(row=r+i, column=c+1).value=var[j]
            i = i + 1
            
def write_lhead(sheet, startcell, var, *args):
    """
    Schreibt die Kopfzeile für eine Excel-Liste
    für eine Variable, die danach mit write_lbody() ausgegeben wird.

    Parameters
    ----------
    sheet : object
        Referenz des zu beschreibenden Worksheets.
    startcell : string
        Erste Zelle der Tabelle. Hier wird die Kopfzeile geschrieben.
    var : string,
        Name der in der Tabelle auszugebenden Variablen.
    *args : string, 
        Namen der Indexmengen der auszugebenden Variablen.

    Returns
    -------
    None.

    """

    r, c = convert_celladdress(startcell)
    # Schreibe Tabellenkopf
    i = 0
    for j in args:
        sheet.cell(row=r, column=c+i).value=j
        i = i + 1
    sheet.cell(row=r, column=c+i).value=var
    
    
def write_lbody(sheet, startcell, var, *args):
    """
    Schreibt die Werte einer Variable, die (praktisch) ungleich Null sind,
    in eine Excel-Tabelle in Form einer Liste

    Parameters
    ----------
    sheet : object
        Referenz des zu beschreibenden Worksheets.
    startcell : string
        Erste Zelle der Tabelle, enthält die Überschrift
        die mit write_lhead() geschrieben wurde.
        Der Tabellenkörper beginnt in der Zeile darunter.
    var : Dictionary mit Tupeln als Keys (Name der auszugebenden Variablen).
    *args : iterable, 
        Indexmengen der Keys des mehrdimensionalen Dictionarys.

    Returns
    -------
    None.

    """

    r, c = convert_celladdress(startcell)
    # Schreibe Rest der Tabelle
    if len(args) > 1:                # falls Indexmenge mehrdimensional ist
        i = 1
        for j in itertools.product(*args):
            if var[j] >= 1e-10 or var[j] <= -1e-10:
                for k in range(0, len(args)):
                    sheet.cell(row=r+i, column=c+k).value=j[k]
                sheet.cell(row=r+i, column=c+k+1).value=var[j]
                i = i + 1 
    else:
        (iterable,) = args           # falls Indexmenge eindimensional ist
        i = 1                       
        for j in iterable:
            if var[j] >= 1e-10 or var[j] <= -1e-10:
                sheet.cell(row=r+i, column=c).value=j
                sheet.cell(row=r+i, column=c+1).value=var[j]
                i = i + 1
                
def out_rc(problem, name, var, *args):
    """
    Printet die reduzierten Kosten für eine Vektorvariable 

    Parameters
    ----------
    problem: Referenz auf Optimierungsinstanz
        Name des Optimierungsproblems
    name: string
        Auszugebender Name der Vektorvariable
    var : Referenz auf Vektorvariableninstanz
        Name der Variable
    *args : Iterable
        Indexmengen der Vektorvariableninstanz

    Returns
    -------
        None.
    """
    print("\n***************************************************")
    print("Reduzierte Kosten der Zielfunktionskoeffizienten von {} \n"
          .format(name))
    if len(args) > 1:
        for i in itertools.product(*args):
            rc = problem.getRedCosts(var[i])
            print("Für {} sind die reduzierten Kosten: {}"
                  .format(var[i], rc))
    else:
        (iterable,) = args
        for i in iterable:
            rc = problem.getRedCosts(var[i])
            print("Für {} sind die reduzierten Kosten: {}"
                  .format(var[i], rc))
            
def out_sl(problem, name, con, *args):
    """
    Printet die Schlupfe für einen Nebenbedingungstyp

    Parameters
    ----------
    problem: Referenz auf Optimierungsinstanz
        Name des Optimierungsproblems
    name: string
        Auszugebender Name des Nebenbedingungstyps
    con : Referenz auf Nebenbedingug
        Name der Nebenbedingung
    *args : Iterable
        Indexmengen des Nebenbedingungstyps

    Returns
    -------
        None.
    """
    print("\n***************************************************")
    print("Schlupfe des Nebenbedingungstyps {} \n"
          .format(name))
    if len(args) > 1:
        slacks = problem.getSlacks(con)
        for j, i in enumerate(itertools.product(*args)):
            print("Für {} {} ist der Schlupf: {}"
                  .format(name, i, slacks[j]))
    else:
        (iterable,) = args
        slacks = problem.getSlacks(con)
        for j, i in enumerate(iterable):
            print("Für {} {} ist der Schlupf: {}"
                  .format(name, i, slacks[j]))

def out_dv(problem, name, con, *args):
    """
    Printet die Werte der Dualvariablen für einen Nebenbedingungstyp

    Parameters
    ----------
    problem: Referenz auf Optimierungsinstanz
        Name des Optimierungsproblems
    name: string
        Auszugebender Name des Nebenbedingungstyps
    con : Referenz auf Nebenbedingug
        Name der Nebenbedingung
    *args : Iterable
        Indexmengen des Nebenbedingungstyps

    Returns
    -------
        None.
    """
    print("\n***************************************************")
    print("Dualwerte des Nebenbedingungstyps {} \n"
          .format(name))
    if len(args) > 1:
        duals = problem.getDuals(con)
        for j, i in enumerate(itertools.product(*args)):
            print("Für {} {} ist der Dualwert: {}"
                  .format(name, i, duals[j]))
    else:
        (iterable,) = args
        duals = problem.getDuals(con)
        for j, i in enumerate(iterable):
            print("Für {} {} ist der Dualwert: {}"
                  .format(name, i, duals[j]))

def saOFC(problem, name, var, *args):
    """
    Printet die ZFK-Intervalle einer Sensitivitätsanalyse
    für eine Vektorvariable 

    Parameters
    ----------
    problem: Referenz auf Optimierungsinstanz
        Name des Optimierungsproblems
    name: string
        Auszugebender Name der Vektorvariable
    var : Referenz auf Vektorvariableninstanz
        Name der Variable
    *args : Iterable
        Indexmengen der Vektorvariableninstanz

    Returns
    -------
        None.
    """
    print("\n***************************************************")
    print("Sensitivitätsanalyse der Zielfunktionskoeffizienten von {} \n"
          .format(name))
    if len(args) > 1:
        for i in itertools.product(*args):
            l, u = problem.objSA([var[i]])
            print("Für {} ist das Intervall: [{}, {}]"
                  .format(var[i], l[0], u[0]))
    else:
        (iterable,) = args
        for i in iterable:
            l, u = problem.objSA([var[i]])
            print("Für {} ist das Intervall: [{}, {}]"
                  .format(var[i], l[0], u[0]))
            
def saRHS(problem, name, con, *args):
    """
    Printet die RHS-Intervalle einer Sensitivitätsanalyse
    für einen Nebenbedingungstyp 

    Parameters
    ----------
    problem: Referenz auf Optimierungsinstanz
        Name des Optimierungsproblems
    name : string
        Auszugebender Name für den Nebenbedingungstyp    
    con : Referenz auf Nebenbedinungstyp
        Name des Nebenbedingungstyps
    *args : Iterable
        Indexmengen des Nebenbedinungungstyps 
        (Achtung! Die Reihenfolge muss simmen!)

    Returns
    -------
        None.
    """
    #print()
    print("\n ***************************************************")
    print("Sensitivitätsanalyse der rechten Seite von {} \n".format(name))
    #print()
    if len(args) > 1:
        rhslower, rhsupper = problem.rhsSA(con)
        for j, i in enumerate(itertools.product(*args)):
            print("Für {} {} ist das Intervall: [{}, {}]"
                  .format(name, i, rhslower[j], rhsupper[j]))
        
    else:
        (iterable,) = args
        rhslower, rhsupper = problem.rhsSA(con)
        for j, i in enumerate(iterable):
            print("Für {} ({}) ist das Intervall: [{}, {}]"
                  .format(name, i, rhslower[j], rhsupper[j]))
            